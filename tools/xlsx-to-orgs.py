#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================
КОНВЕРТЕР СПРАВОЧНИКА ОРГАНИЗАЦИЙ:  Excel (.xlsx) -> данные сайта
================================================================
ЗАЧЕМ. Оргкомитет ведёт справочник образовательных организаций
в Excel (колонки FullName / ShortName / RegionName – см. файл
docs-dev/reference/orgs-source.xlsx). Сайт не читает .xlsx
напрямую: этот скрипт превращает выгрузку в лёгкие JSON-файлы,
которые страница регистрации подгружает ПО ВЫБРАННОМУ РЕГИОНУ
(assets/js/reg-form.js), не качая всю базу целиком.

КАК ЗАПУСТИТЬ (когда пришлют новую версию справочника):
  1) положить новый .xlsx в docs-dev/reference/orgs-source.xlsx
  2) pip install openpyxl
  3) python3 tools/xlsx-to-orgs.py
  4) закоммитить изменённые файлы assets/data/orgs/

ЧТО ПОЛУЧАЕТСЯ:
  assets/data/orgs/r01.json … r89.json   - организации каждого
      из 89 субъектов РФ (нумерация = порядок регионов в списке
      REGIONS внутри assets/js/reg-form.js; файл rNN соответствует
      REGIONS[NN-1]);
  assets/data/orgs/none.json             - записи без региона
      (в исходной выгрузке поле пустое; это обычные школы-
      филиалы, их подмешиваем в поиск любого региона);
  assets/data/orgs/foreign.json          - организации за
      пределами РФ (в форме сейчас только 89 субъектов РФ,
      поэтому в поиске не участвуют; файл сохранён на будущее);
  assets/data/orgs/manifest.json         - служебная сводка:
      дата сборки, счётчики, какой файл за какой регион отвечает.

ФОРМАТ ФАЙЛА РЕГИОНА: {"items": [[FullName, ShortName], ...]}
Пустой ShortName записывается как "". Кодировка UTF-8 (кириллица
как есть) - браузер понимает сам, сжатие на сервере догоняет.
================================================================
"""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Нужен пакет openpyxl:  pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "docs-dev" / "reference" / "orgs-source.xlsx"
OUT = ROOT / "assets" / "data" / "orgs"

# 89 субъектов РФ. ПОРЯДОК КРИТИЧЕН: он должен совпадать
# со списком REGIONS в assets/js/reg-form.js, потому что имя файла
# rNN.json вычисляется по номеру региона в этом списке.
REGIONS = ["Республика Адыгея", "Республика Алтай", "Республика Башкортостан", "Республика Бурятия",
    "Республика Дагестан", "Донецкая Народная Республика", "Республика Ингушетия", "Кабардино-Балкарская Республика",
    "Республика Калмыкия", "Карачаево-Черкесская Республика", "Республика Карелия", "Республика Коми",
    "Республика Крым", "Луганская Народная Республика", "Республика Марий Эл", "Республика Мордовия",
    "Республика Саха (Якутия)", "Республика Северная Осетия – Алания", "Республика Татарстан", "Республика Тыва",
    "Удмуртская Республика", "Республика Хакасия", "Чеченская Республика", "Чувашская Республика",
    "Алтайский край", "Забайкальский край", "Камчатский край", "Краснодарский край", "Красноярский край",
    "Пермский край", "Приморский край", "Ставропольский край", "Хабаровский край",
    "Амурская область", "Архангельская область", "Астраханская область", "Белгородская область", "Брянская область",
    "Владимирская область", "Волгоградская область", "Вологодская область", "Воронежская область", "Запорожская область",
    "Ивановская область", "Иркутская область", "Калининградская область", "Калужская область", "Кемеровская область – Кузбасс",
    "Кировская область", "Костромская область", "Курганская область", "Курская область", "Ленинградская область",
    "Липецкая область", "Магаданская область", "Московская область", "Мурманская область", "Нижегородская область",
    "Новгородская область", "Новосибирская область", "Омская область", "Оренбургская область", "Орловская область",
    "Пензенская область", "Псковская область", "Ростовская область", "Рязанская область", "Самарская область",
    "Саратовская область", "Сахалинская область", "Свердловская область", "Смоленская область", "Тамбовская область",
    "Тверская область", "Томская область", "Тульская область", "Тюменская область", "Ульяновская область",
    "Херсонская область", "Челябинская область", "Ярославская область",
    "Москва", "Санкт-Петербург", "Севастополь",
    "Еврейская автономная область", "Ненецкий автономный округ", "Ханты-Мансийский автономный округ – Югра",
    "Чукотский автономный округ", "Ямало-Ненецкий автономный округ"]

# Как называния регионов в выгрузке Excel приводятся к нашему списку.
REGION_ALIASES = {
    "г. Москва": "Москва",
    "г. Санкт-Петербург": "Санкт-Петербург",
    "г. Севастополь": "Севастополь",
    "Республика Адыгея (Адыгея)": "Республика Адыгея",
    "Республика Северная Осетия - Алания": "Республика Северная Осетия – Алания",
    "Республика Татарстан (Татарстан)": "Республика Татарстан",
    "Чувашская Республика - Чувашия": "Чувашская Республика",
    "Кемеровская область": "Кемеровская область – Кузбасс",
    "Ханты-Мансийский автономный округ - Югра": "Ханты-Мансийский автономный округ – Югра",
}
FOREIGN_LABEL = "образовательные учреждения, находящиеся за пределами Российской Федерации"

# Служебные артефакты Excel: "_x000D_" (escaped CR), управляющие символы.
JUNK_RE = re.compile(r"_x[0-9A-Fa-f]{4}_|[\x00-\x08\x0b\x0c\x0e-\x1f]")
SPACE_RE = re.compile(r"\s+")


def clean(text):
    """Убрать служебный мусор Excel, схлопнуть пробелы. Состав
    названия (кавычки-ёлочки/лапки, регистр) не трогаем: поиск
    должен совпадать с официальным справочником дословно."""
    if text is None:
        return ""
    return SPACE_RE.sub(" ", JUNK_RE.sub(" ", str(text))).strip()


def canon_region(raw):
    """RegionName из выгрузки -> (ключ файла, каноническое имя).
    Ключи: 'r01'...'r89', 'none', 'foreign'."""
    r = clean(raw)
    if not r or r.lower() == "none":
        return "none", ""
    r = REGION_ALIASES.get(r, r)
    if r == FOREIGN_LABEL or "за пределами" in r:
        return "foreign", ""
    if r in REGIONS:
        return "r%02d" % (REGIONS.index(r) + 1), r
    sys.exit(f"Неизвестный регион в выгрузке: {raw!r}. "
             f"Добавьте его в REGION_ALIASES в этом скрипте.")


def main():
    if not SRC.exists():
        sys.exit(f"Не найден исходник: {SRC}")
    print(f"Читаю {SRC.name} …")
    wb = openpyxl.load_workbook(SRC, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = [clean(c) for c in next(rows)]
    try:
        c_full = header.index("FullName")
        c_short = header.index("ShortName")
        c_reg = header.index("RegionName")
    except ValueError:
        sys.exit(f"В первой строке жду колонки FullName / ShortName / "
                 f"RegionName, а вижу: {header}")

    buckets = {}  # ключ -> {(full, short)}
    skipped = 0
    for row in rows:
        full = clean(row[c_full])
        if not full:
            skipped += 1
            continue
        short = clean(row[c_short])
        if short in ("-", "None"):
            short = ""
        key, _ = canon_region(row[c_reg])
        buckets.setdefault(key, set()).add((full, short))

    OUT.mkdir(parents=True, exist_ok=True)
    manifest = {"version": 1,
                "source": "docs-dev/reference/orgs-source.xlsx",
                "note": "rNN.json = REGIONS[NN-1] из assets/js/reg-form.js; "
                        "none.json подмешивается в поиск любого региона",
                "regions": {}, "files": {}}
    total = 0
    for key, pairs in sorted(buckets.items()):
        items = sorted(pairs, key=lambda p: p[0].lower())
        fname = key + ".json"
        with open(OUT / fname, "w", encoding="utf-8") as f:
            json.dump({"items": [[n, s] for n, s in items]}, f,
                      ensure_ascii=False, separators=(",", ":"))
        total += len(items)
        manifest["files"][fname] = len(items)
        if key.startswith("r"):
            rr = REGIONS[int(key[1:]) - 1]
            manifest["regions"][rr] = fname
    with open(OUT / "manifest.json", "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=1)

    print(f"Готово: {total} организаций (пропущено пустых строк: {skipped})")
    print(f"Файлов: {len(buckets)} -> {OUT.relative_to(ROOT)}")
    for fname, cnt in sorted(manifest['files'].items(),
                             key=lambda kv: -kv[1])[:5]:
        print(f"  самый крупный: {fname}: {cnt}")
    missing = [r for r in REGIONS if r not in manifest["regions"]]
    if missing:
        print("ВНИМАНИЕ: регионы без единой организации:", missing)


if __name__ == "__main__":
    main()
